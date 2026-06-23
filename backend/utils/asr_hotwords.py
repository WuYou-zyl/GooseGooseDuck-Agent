from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


HOTWORD_OUTPUT = Path("data/rule_library/asr_hotwords.txt")
RULE_LIBRARY_DIR = Path("data/rule_library")


def _clean_token(token: str) -> str:
    token = token.replace("\u200b", "").replace("\ufeff", "").strip()
    token = re.sub(r"\s+", "", token)
    token = token.strip("：:;；，,。.!?？、()（）[]【】\"'“”‘’")
    return token


def _is_valid_hotword(token: str) -> bool:
    if not token:
        return False
    if len(token) < 2:
        return False
    if token.isdigit():
        return False
    if any(ch in token for ch in ("\n", "\r", "\t")):
        return False
    return True


def normalize_hotword(token: str) -> str:
    token = _clean_token(str(token))
    return token if _is_valid_hotword(token) else ""


def _split_candidates(text: str) -> list[str]:
    raw_parts = re.split(r"[：:;/／,，、。.!?？\s]+", text)
    out: list[str] = []
    for part in raw_parts:
        token = _clean_token(part)
        if _is_valid_hotword(token):
            out.append(token)
    return out


def _collect_from_txt(path: Path) -> set[str]:
    words: set[str] = set()
    content = path.read_text(encoding="utf-8", errors="ignore")
    for line in content.splitlines():
        line = line.strip()
        if not line:
            continue
        if "：" in line:
            head = line.split("：", 1)[0]
            for part in re.split(r"[/／]", head):
                token = _clean_token(part)
                if _is_valid_hotword(token):
                    words.add(token)
        for token in _split_candidates(line):
            if 2 <= len(token) <= 8:
                words.add(token)
    return words


def _collect_from_xlsx(path: Path) -> set[str]:
    words: set[str] = set()
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not row:
                continue
            first = row[0]
            if first is None:
                continue
            token = _clean_token(str(first))
            if row_idx == 1 and token in {"角色名称", "地图名称", "名称"}:
                continue
            if _is_valid_hotword(token):
                words.add(token)
    return words


def collect_hotwords(rule_library_dir: Path | None = None) -> list[str]:
    base = rule_library_dir or RULE_LIBRARY_DIR
    words: set[str] = set()

    for path in sorted(base.glob("*.txt")):
        words.update(_collect_from_txt(path))

    for path in sorted(base.glob("*.xlsx")):
        words.update(_collect_from_xlsx(path))

    filtered = [
        w
        for w in words
        if 2 <= len(w) <= 12
        and not w.startswith("http")
        and w not in {"玩法推荐", "玩法特点", "游玩技巧", "核心技能", "胜利条件"}
    ]
    return sorted(set(filtered), key=lambda x: (len(x), x))


def ensure_hotword_file(output_path: Path | None = None) -> Path:
    output = output_path or HOTWORD_OUTPUT
    hotwords = collect_hotwords()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(hotwords) + "\n", encoding="utf-8")
    return output


def load_hotwords_text(path: Path | None = None) -> str:
    output = path or HOTWORD_OUTPUT
    if not output.exists():
        ensure_hotword_file(output)
    return output.read_text(encoding="utf-8", errors="ignore").strip()


def merge_hotwords_text(base_text: str, extra_words: list[str] | tuple[str, ...] | None) -> str:
    seen: set[str] = set()
    merged: list[str] = []

    def _push(word: str) -> None:
        token = normalize_hotword(word)
        if not token or token in seen:
            return
        seen.add(token)
        merged.append(token)

    for line in (base_text or "").splitlines():
        _push(line)

    for word in extra_words or []:
        _push(word)

    return "\n".join(merged)
